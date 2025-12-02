"""
Financial Excel Query Engine V5.2 (Universal Note Filtering)
------------------------------------------------------------
Enhanced with aggressive, universal footnote/note filtering system.

New Features:
- Multi-signal note detection (spatial + pattern + semantic + positional)
- Aggressive isolation detection with connectivity analysis
- Bottom-of-sheet note zone detection
- Semantic note detection using language patterns
"""

import os
import re
import warnings
import numpy as np
import pandas as pd
from typing import List, Dict, Tuple, Any, Optional, Literal, Set
from dataclasses import dataclass, field
from enum import Enum
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from tenacity import retry, stop_after_attempt, wait_exponential

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ==========================================
# 1. Configuration & Data Structures
# ==========================================

@dataclass
class QueryEngineConfig:
    """Configuration for the engine."""
    semantic_backend: Literal['basic', 'openai'] = 'openai'
    
    openai_api_key: Optional[str] = field(default_factory=lambda: os.getenv("OPENAI_API_KEY"))
    openai_model: str = "text-embedding-3-small"
    
    min_confidence: float = 0.5
    exact_match_boost: float = 0.25
    
    header_text_ratio_threshold: float = 0.55
    style_weight_boost: float = 0.35
    use_style_analysis: bool = True
    max_header_rows: int = 10
    max_empty_rows_between_tables: int = 3
    min_numeric_cells_for_data_row: int = 2
    
    # Enhanced Footnote Filtering Settings
    enable_footnote_filtering: bool = True
    aggressive_note_filtering: bool = True  # More strict filtering
    min_distance_from_ int = 2  # Reduced threshold for more aggressive filtering
    isolation_window: int = 3  # Check 3x3 cells around for connectivity
    sparse_region_threshold: float = 0.2  # Lower = more aggressive (20% filled)
    bottom_zone_threshold: float = 0.15  # Bottom 15% of sheet = note zone
    
    # Pattern matching for notes
    footnote_patterns: List[str] = field(default_factory=lambda: [
        r'^\*+',
        r'^\d+\)',
        r'^\[\d+\]',
        r'^note[s]?\s*:',
        r'^source[s]?\s*:',
        r'^see\s+',
        r'^disclaimer\s*:',
        r'^assumption[s]?\s*:',
        r'^\(\d+\)',
        r'^[a-z]\)',
        r'^\d+\.\s',
        r'^ref[s]?\s*:',
        r'^legend\s*:',
        r'^important\s*:',
        r'^warning\s*:',
        r'^\([a-z]\)',
    ])
    
    # Semantic patterns (word-level detection)
    note_keywords: Set[str] = field(default_factory=lambda: {
        'note', 'notes', 'footnote', 'source', 'sources', 'disclaimer',
        'assumption', 'assumptions', 'reference', 'references', 'see',
        'legend', 'annotation', 'remark', 'comment', 'explanation',
        'methodology', 'definition', 'prepared by', 'compiled by',
        'as of', 'unaudited', 'preliminary', 'subject to change'
    })
    
    max_footnote_length: int = 500

@dataclass
class SearchResult:
    """Standardized result object."""
    value: Any
    confidence: float
    sheet_name: str
    row: int
    col: int
    header_path: List[str]
    context_text: str
    value_type: str
    is_exact_match: bool = False

    def __repr__(self):
        path = " > ".join(self.header_path)
        exact_marker = " [EXACT]" if self.is_exact_match else ""
        return (f"<Result value={self.value} | confidence={self.confidence:.2f}{exact_marker} | "
                f"loc={self.sheet_name}!R{self.row+1}C{self.col+1} | path='{path}'>")

class CellType(Enum):
    HEADER_HORIZONTAL = "header_horizontal"
    HEADER_VERTICAL = "header_vertical"
    DATA_VALUE = "data_value"
    INDEX = "index"
    METADATA = "metadata"
    FOOTNOTE = "footnote"
    EMPTY = "empty"

class ValueType(Enum):
    NUMBER = "number"
    TEXT = "text"
    PERCENTAGE = "percentage"
    DATE = "date"
    CURRENCY = "currency"
    ANY = "any"

# ==========================================
# 2. Universal Footnote/Note Detection
# ==========================================

class UniversalNoteFilter:
    """
    Aggressive, multi-signal note/footnote detector.
    Uses: Pattern matching + Spatial analysis + Connectivity + Semantic detection + Positional heuristics
    """
    
    def __init__(self, config: QueryEngineConfig):
        self.config = config
        self.compiled_patterns = [re.compile(pattern, re.IGNORECASE) for pattern in config.footnote_patterns]
        self.note_keywords = {kw.lower() for kw in config.note_keywords}
    
    def is_note_by_pattern(self, value: Any) -> bool:
        """Check if cell matches note/footnote patterns."""
        if not isinstance(value, str):
            return False
        
        value_stripped = value.strip()
        
        if len(value_stripped) < 2:
            return False
        
        if len(value_stripped) > self.config.max_footnote_length:
            return False
        
        # Check regex patterns
        for pattern in self.compiled_patterns:
            if pattern.match(value_stripped):
                return True
        
        return False
    
    def is_note_by_semantics(self, value: Any) -> bool:
        """Check if cell contains note-like keywords."""
        if not isinstance(value, str):
            return False
        
        value_lower = value.lower().strip()
        
        # Check if entire value is a note keyword
        if value_lower in self.note_keywords:
            return True
        
        # Check if value starts with note keywords
        words = value_lower.split()
        if len(words) > 0 and words[0] in self.note_keywords:
            return True
        
        # Check for phrases
        for keyword in self.note_keywords:
            if len(keyword.split()) > 1:  # Multi-word keywords
                if keyword in value_lower:
                    return True
        
        return False
    
    def detect_data_boundaries(self, df: pd.DataFrame) -> Dict[str, int]:
        """Find the bounding box of actual data content."""
        non_empty_rows = [i for i in range(len(df)) if not df.iloc[i].isna().all()]
        
        if not non_empty_rows:
            return {'min_row': 0, 'max_row': 0, 'min_col': 0, 'max_col': 0}
        
        min_row = min(non_empty_rows)
        max_row = max(non_empty_rows)
        
        non_empty_cols = [i for i in range(len(df.columns)) if not df.iloc[:, i].isna().all()]
        
        min_col = min(non_empty_cols) if non_empty_cols else 0
        max_col = max(non_empty_cols) if non_empty_cols else 0
        
        return {
            'min_row': min_row,
            'max_row': max_row,
            'min_col': min_col,
            'max_col': max_col
        }
    
    def is_in_note_zone(self, row_idx: int, col_idx: int, 
                        boundaries: Dict[str, int], df: pd.DataFrame) -> bool:
        """
        Check if cell is in typical 'note zone' locations:
        - Bottom 15% of sheet with sparse data
        - Far right column with sparse data
        - Isolated single cells below/beside tables
        """
        max_row = len(df)
        max_col = len(df.columns)
        
        # Check if in bottom zone
        bottom_threshold = max_row * (1 - self.config.bottom_zone_threshold)
        if row_idx > bottom_threshold and row_idx > boundaries['max_row']:
            return True
        
        # Check if far below the last data row
        if row_idx > boundaries['max_row'] + self.config.min_distance_from_
            return True
        
        # Check if far to the right of last data column
        if col_idx > boundaries['max_col'] + self.config.min_distance_from_
            return True
        
        return False
    
    def is_connected_to_table(self, row_idx: int, col_idx: int, 
                             df: pd.DataFrame, tables: List[Dict]) -> bool:
        """
        Check if cell is part of or adjacent to any detected table.
        A cell is connected if it's within the table boundaries or immediately adjacent.
        """
        for table in tables:
            h_start, h_end = table['header_range']
            d_start, d_end = table['data_range']
            v_cols = table.get('vertical_header_cols', [0])
            
            # Get table column range
            # Assume table spans from first vertical header to last data column
            table_min_col = min(v_cols) if v_cols else 0
            
            # Find last non-empty column in this table's data range
            table_max_col = 0
            for r in range(d_start, d_end):
                row_data = df.iloc[r]
                last_valid = row_data.last_valid_index()
                if last_valid is not None:
                    table_max_col = max(table_max_col, last_valid)
            
            # Check if cell is within or immediately adjacent to table
            row_in_range = (h_start - 1) <= row_idx <= (d_end + 1)
            col_in_range = (table_min_col - 1) <= col_idx <= (table_max_col + 1)
            
            if row_in_range and col_in_range:
                return True
        
        return False
    
    def is_spatially_isolated(self, row_idx: int, col_idx: int, 
                             df: pd.DataFrame) -> bool:
        """
        Aggressive spatial isolation check using connectivity analysis.
        A cell is isolated if it's surrounded by mostly empty cells.
        """
        window = self.config.isolation_window
        
        row_start = max(0, row_idx - window)
        row_end = min(len(df), row_idx + window + 1)
        col_start = max(0, col_idx - window)
        col_end = min(len(df.columns), col_idx + window + 1)
        
        region = df.iloc[row_start:row_end, col_start:col_end]
        non_empty_count = region.notna().sum().sum()
        total_cells = region.size
        
        # Calculate density
        density = non_empty_count / total_cells if total_cells > 0 else 0
        
        # Isolated if density is below threshold
        return density < self.config.sparse_region_threshold
    
    def is_single_cell_row(self, row_idx: int, df: pd.DataFrame) -> bool:
        """Check if row contains only one non-empty cell (likely a note)."""
        row = df.iloc[row_idx]
        non_empty = row.dropna()
        return len(non_empty) == 1
    
    def is_different_formatting_cluster(self, row_idx: int, col_idx: int,
                                       style_grid: List[List[Dict]],
                                       table_styles: Dict) -> bool:
        """
        Check if cell has significantly different formatting from the table.
        Notes often have smaller font, italic, or different alignment.
        """
        if row_idx >= len(style_grid) or col_idx >= len(style_grid[row_idx]):
            return False
        
        cell_style = style_grid[row_idx][col_idx]
        
        # Compare font size
        if cell_style.get('font_size', 11) < table_styles.get('avg_font_size', 11) - 1:
            return True
        
        return False
    
    def should_filter_cell(self, value: Any, row_idx: int, col_idx: int,
                          boundaries: Dict[str, int], df: pd.DataFrame,
                          tables: List[Dict], style_grid: List[List[Dict]] = None) -> Tuple[bool, str]:
        """
        Universal note detection using multiple signals.
        Returns: (should_filter, reason)
        """
        
        if not self.config.enable_footnote_filtering:
            return False, ""
        
        # Signal 1: Pattern matching (highest priority)
        if self.is_note_by_pattern(value):
            return True, "pattern_match"
        
        # Signal 2: Semantic keyword detection
        if isinstance(value, str) and self.is_note_by_semantics(value):
            return True, "semantic_keyword"
        
        if not self.config.aggressive_note_filtering:
            return False, ""
        
        # === AGGRESSIVE FILTERING SIGNALS ===
        
        # Signal 3: In typical note zone
        if self.is_in_note_zone(row_idx, col_idx, boundaries, df):
            # Additional check: if it's text in note zone, filter it
            if isinstance(value, str):
                return True, "note_zone"
        
        # Signal 4: Not connected to any table
        if not self.is_connected_to_table(row_idx, col_idx, df, tables):
            # If it's text and not connected, likely a note
            if isinstance(value, str):
                return True, "disconnected_from_table"
        
        # Signal 5: Spatial isolation
        if self.is_spatially_isolated(row_idx, col_idx, df):
            # If isolated text cell, filter it
            if isinstance(value, str):
                return True, "spatial_isolation"
        
        # Signal 6: Single cell in entire row (common for notes)
        if self.is_single_cell_row(row_idx, df):
            if isinstance(value, str):
                # Check if it's far from data
                if row_idx > boundaries['max_row'] + 1 or row_idx < boundaries['min_row'] - 1:
                    return True, "isolated_single_cell_row"
        
        return False, ""

# ==========================================
# 3. Semantic Matching Backends
# ==========================================

class SemanticMatcher:
    """Base interface for similarity."""
    def calculate_similarity(self, query: str, target: str) -> float:
        raise NotImplementedError

class BasicSemanticMatcher(SemanticMatcher):
    """Fallback matcher using token overlap."""
    def normalize(self, text: str) -> set:
        text = str(text).lower().strip()
        text = re.sub(r'[^\w\s]', '', text)
        return set(text.split())

    def calculate_similarity(self, query: str, target: str) -> float:
        q_lower = query.lower()
        t_lower = target.lower()
        
        if q_lower in t_lower or t_lower in q_lower:
            return 0.95
        
        q_tok = self.normalize(query)
        t_tok = self.normalize(target)
        if not q_tok or not t_tok: return 0.0
        
        intersection = len(q_tok & t_tok)
        union = len(q_tok | t_tok)
        
        jaccard = intersection / union if union > 0 else 0.0
        coverage = intersection / len(q_tok) if q_tok else 0.0
        
        return max(jaccard, coverage * 0.9)

class OpenAISemanticMatcher(SemanticMatcher):
    """Enterprise matcher using OpenAI Embeddings."""
    def __init__(self, api_key: str, model: str):
        try:
            from openai import OpenAI
            self.client = OpenAI(api_key=api_key)
            self.model = model
        except ImportError:
            raise ImportError("Please install openai: `pip install openai`")
        except Exception as e:
            raise ValueError(f"Failed to init OpenAI client: {e}")

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
    def embed_batch(self, texts: List[str], batch_size: int = 100) -> List[List[float]]:
        if not texts: return []
        clean_texts = [t.replace("\n", " ")[:8000] for t in texts]
        
        response = self.client.embeddings.create(
            input=clean_texts,
            model=self.model
        )
        return [data.embedding for data in response.data]

    def cosine_similarity(self, vec_a: List[float], vec_b: List[float]) -> float:
        a = np.array(vec_a)
        b = np.array(vec_b)
        norm_a = np.linalg.norm(a)
        norm_b = np.linalg.norm(b)
        if norm_a == 0 or norm_b == 0: return 0.0
        return float(np.dot(a, b) / (norm_a * norm_b))

# ==========================================
# 4. Structure & Data Handling
# ==========================================

class MergedCellHandler:
    """Handles Excel merged cells and style extraction."""
    def __init__(self, file_path: str):
        self.wb = load_workbook(file_path, data_only=True)
        self.merged_map = {}
        
        for sheet in self.wb.worksheets:
            for group in sheet.merged_cells.ranges:
                min_row, min_col = group.min_row, group.min_col
                for r in range(min_row, group.max_row + 1):
                    for c in range(min_col, group.max_col + 1):
                        self.merged_map[(sheet.title, r, c)] = (min_row, min_col)

    def get_sheet_data_and_styles(self, sheet_name: str) -> Tuple[pd.DataFrame, List[List[Dict]]]:
        """Returns dataframe and style grid."""
        sheet = self.wb[sheet_name]
        data_rows = []
        style_grid = []

        for r_idx, row in enumerate(sheet.iter_rows(), start=1):
            row_data = []
            row_styles = []
            for c_idx, cell in enumerate(row, start=1):
                lookup_r, lookup_c = self.merged_map.get((sheet_name, r_idx, c_idx), (r_idx, c_idx))
                real_cell = sheet.cell(row=lookup_r, column=lookup_c)
                
                row_data.append(real_cell.value)
                
                style_info = {
                    'bold': bool(real_cell.font and real_cell.font.bold),
                    'font_size': real_cell.font.size if real_cell.font else 11,
                    'alignment': real_cell.alignment.horizontal if real_cell.alignment else None,
                    'is_merged': (sheet_name, r_idx, c_idx) in self.merged_map,
                    'italic': bool(real_cell.font and real_cell.font.italic)
                }
                row_styles.append(style_info)
            
            data_rows.append(row_data)
            style_grid.append(row_styles)

        return pd.DataFrame(data_rows), style_grid

class CellClassifier:
    """Classifies cells as headers vs data."""
    def __init__(self, config: QueryEngineConfig):
        self.config = config
    
    def classify_cell(self, value: Any, style: Dict, row_idx: int, 
                     col_idx: int, row_ pd.Series, col_ pd.Series) -> CellType:
        """Classifies a single cell."""
        if pd.isna(value) or str(value).strip() == "":
            return CellType.EMPTY
        
        is_bold = style.get('bold', False)
        font_size = style.get('font_size', 11)
        is_merged = style.get('is_merged', False)
        alignment = style.get('alignment', None)
        
        is_text = isinstance(value, str)
        is_numeric = isinstance(value, (int, float)) and not isinstance(value, bool)
        
        row_has_mostly_text = self._row_has_mostly_text(row_data)
        col_has_mostly_numbers = self._col_has_mostly_numbers(col_data)
        
        header_score = 0.0
        if is_bold: header_score += 0.3
        if font_size > 11: header_score += 0.2
        if is_merged: header_score += 0.15
        if is_text: header_score += 0.1
        if row_has_mostly_text: header_score += 0.15
        if alignment == 'center': header_score += 0.1
        
        if header_score >= 0.5:
            if col_idx <= 2 and not col_has_mostly_numbers:
                return CellType.HEADER_VERTICAL
            return CellType.HEADER_HORIZONTAL
        
        if is_numeric and col_has_mostly_numbers:
            return CellType.DATA_VALUE
        
        if is_text and col_idx == 0:
            return CellType.INDEX
        
        return CellType.DATA_VALUE
    
    def _row_has_mostly_text(self, row: pd.Series) -> bool:
        clean = row.dropna()
        if len(clean) == 0: return False
        text_count = sum(isinstance(x, str) for x in clean)
        return (text_count / len(clean)) >= 0.6
    
    def _col_has_mostly_numbers(self, col: pd.Series) -> bool:
        clean = col.dropna()
        if len(clean) == 0: return False
        num_count = sum(isinstance(x, (int, float)) and not isinstance(x, bool) for x in clean)
        return (num_count / len(clean)) >= 0.6

class AdvancedStructureDetector:
    """Enhanced table detection."""
    def __init__(self, config: QueryEngineConfig):
        self.config = config
        self.classifier = CellClassifier(config)

    def detect_tables(self, df: pd.DataFrame, style_grid: List[List[Dict]]) -> List[Dict]:
        """Detects multiple tables on a sheet."""
        tables = []
        i = 0
        max_rows = len(df)

        while i < max_rows:
            if df.iloc[i].isna().all():
                i += 1
                continue

            header_start = i
            header_end = self._find_header_end(df, style_grid, i)
            
            if header_end == header_start:
                i += 1
                continue
            
            data_start = header_end
            data_end = self._find_data_end(df, data_start)
            
            if data_end > data_start:
                vertical_header_cols = self._detect_vertical_headers(
                    df.iloc[data_start:data_end], 
                    [style_grid[r] for r in range(data_start, data_end)]
                )
                
                tables.append({
                    'header_range': (header_start, header_end),
                    'data_range': (data_start, data_end),
                    'vertical_header_cols': vertical_header_cols,
                    'name': self._infer_table_name(df, header_start, header_end)
                })
                i = data_end
            else:
                i += 1
        
        return tables

    def _find_header_end(self, df: pd.DataFrame, style_grid: List[List[Dict]], start: int) -> int:
        end = start
        max_rows = min(start + self.config.max_header_rows, len(df))
        
        for r in range(start, max_rows):
            row_vals = df.iloc[r]
            row_styles = style_grid[r] if r < len(style_grid) else []
            
            if self._is_header_row(row_vals, row_styles, df):
                end = r + 1
            else:
                if self._is_data_row(row_vals):
                    break
                elif row_vals.isna().all():
                    continue
                else:
                    end = r + 1
        
        return end
    
    def _is_header_row(self, row: pd.Series, styles: List[Dict], df: pd.DataFrame) -> bool:
        clean_row = row.dropna()
        if len(clean_row) == 0: return False
        
        text_count = sum(isinstance(x, str) for x in clean_row)
        text_ratio = text_count / len(clean_row)
        
        bold_count = sum(1 for idx, style in enumerate(styles) 
                        if idx < len(row) and pd.notna(row.iloc[idx]) and style.get('bold', False))
        bold_ratio = bold_count / len(clean_row) if len(clean_row) > 0 else 0
        
        merged_count = sum(1 for style in styles if style.get('is_merged', False))
        merged_ratio = merged_count / len(styles) if len(styles) > 0 else 0
        
        avg_font_size = np.mean([s.get('font_size', 11) for s in styles])
        font_boost = 0.1 if avg_font_size > 11 else 0
        
        header_score = (text_ratio * 0.4) + (bold_ratio * 0.3) + (merged_ratio * 0.2) + font_boost
        
        return header_score >= self.config.header_text_ratio_threshold
    
    def _is_data_row(self, row: pd.Series) -> bool:
        clean_row = row.dropna()
        if len(clean_row) == 0: return False
        
        num_count = sum(isinstance(x, (int, float)) and not isinstance(x, bool) for x in clean_row)
        return num_count >= self.config.min_numeric_cells_for_data_row
    
    def _find_data_end(self, df: pd.DataFrame, start: int) -> int:
        end = start
        empty_counter = 0
        max_rows = len(df)
        
        for r in range(start, max_rows):
            if df.iloc[r].isna().all():
                empty_counter += 1
                if empty_counter >= self.config.max_empty_rows_between_tables:
                    break
            else:
                empty_counter = 0
                end = r + 1
        
        return end
    
    def _detect_vertical_headers(self, data_block: pd.DataFrame, 
                                 style_block: List[List[Dict]]) -> List[int]:
        vertical_cols = []
        
        for col_idx in range(min(3, len(data_block.columns))):
            col_data = data_block.iloc[:, col_idx]
            col_styles = [row[col_idx] if col_idx < len(row) else {} for row in style_block]
            
            text_ratio = sum(isinstance(x, str) for x in col_data.dropna()) / max(len(col_data.dropna()), 1)
            bold_count = sum(1 for s in col_styles if s.get('bold', False))
            bold_ratio = bold_count / max(len(col_styles), 1)
            has_indentation = any(isinstance(v, str) and (v.startswith('  ') or v.startswith('\t')) 
                                 for v in col_data if pd.notna(v))
            
            if text_ratio >= 0.7 or (text_ratio >= 0.5 and bold_ratio >= 0.3) or has_indentation:
                vertical_cols.append(col_idx)
        
        return vertical_cols if vertical_cols else [0]
    
    def _infer_table_name(self, df: pd.DataFrame, header_start: int, header_end: int) -> str:
        for r in range(max(0, header_start - 3), header_start):
            row = df.iloc[r]
            non_empty = row.dropna()
            if len(non_empty) == 1:
                return str(non_empty.iloc[0])
        
        first_header = df.iloc[header_start].dropna()
        return str(first_header.iloc[0]) if len(first_header) > 0 else "Table"

# ==========================================
# 5. Main Engine
# ==========================================

class FinancialExcelEngine:
    """Production-grade Financial Excel Query Engine with Universal Note Filtering."""
    def __init__(self, file_path: str, config: QueryEngineConfig):
        self.file_path = file_path
        self.config = config
        self.records = []
        self.vector_index = {}
        self.filtered_notes = []
        self.tables_by_sheet = {}
        
        self.merge_handler = MergedCellHandler(file_path)
        self.detector = AdvancedStructureDetector(config)
        self.note_filter = UniversalNoteFilter(config)
        
        if config.semantic_backend == 'openai':
            if not config.openai_api_key:
                raise ValueError("OpenAI API Key missing.")
            self.matcher = OpenAISemanticMatcher(config.openai_api_key, config.openai_model)
        else:
            self.matcher = BasicSemanticMatcher()

        print("🔍 Parsing file structure...")
        self._ingest_file()
        
        if config.semantic_backend == 'openai':
            print("🧠 Generating embeddings...")
            self._build_embeddings()
        
        print(f"✅ Engine Ready. Indexed {len(self.records)} data points.")
        if self.config.enable_footnote_filtering:
            print(f"🗑️  Filtered {len(self.filtered_notes)} note/footnote cells.")

    def _ingest_file(self):
        """Parse all sheets with aggressive note filtering."""
        wb = load_workbook(self.file_path, read_only=True)
        
        for sheet_name in wb.sheetnames:
            df, style_grid = self.merge_handler.get_sheet_data_and_styles(sheet_name)
            
            # Detect data boundaries
            boundaries = self.note_filter.detect_data_boundaries(df)
            
            # Detect tables
            tables = self.detector.detect_tables(df, style_grid)
            self.tables_by_sheet[sheet_name] = tables
            
            print(f"  📊 {sheet_name}: Found {len(tables)} table(s)")
            
            # Process each table
            for tbl_idx, tbl in enumerate(tables):
                self._process_table(df, style_grid, sheet_name, tbl, tbl_idx, boundaries, tables)

    def _process_table(self, df: pd.DataFrame, style_grid: List[List[Dict]], 
                      sheet_name: str, table: Dict, table_idx: int, 
                      boundaries: Dict[str, int], all_tables: List[Dict]):
        """Process table with note filtering in BOTH headers and values."""
        h_start, h_end = table['header_range']
        d_start, d_end = table['data_range']
        v_header_cols = table['vertical_header_cols']
        
        header_block = df.iloc[h_start:h_end]
        
        # 🔥 Pass additional context for note filtering
        col_paths = self._build_column_paths(
            header_block, 
            style_grid[h_start:h_end] if style_grid else None,
            boundaries,
            df,
            all_tables
        )
        
        for r_idx in range(d_start, d_end):
            row_data = df.iloc[r_idx]
            
            # 🔥 Filter notes from row headers too
            row_path = self._build_row_path(
                row_data, 
                v_header_cols,
                r_idx,
                boundaries,
                df,
                all_tables,
                style_grid,
                sheet_name
            )
            
            if not row_path:
                continue
            
            for c_idx, val in enumerate(row_data):
                if c_idx in v_header_cols or pd.isna(val) or str(val).strip() == "":
                    continue
                
                # Existing value filtering
                should_filter, reason = self.note_filter.should_filter_cell(
                    val, r_idx, c_idx, boundaries, df, all_tables, style_grid
                )
                
                if should_filter:
                    self.filtered_notes.append({
                        'sheet': sheet_name,
                        'row': r_idx,
                        'col': c_idx,
                        'value': val,
                        'reason': reason
                    })
                    continue
                
                full_path = row_path + col_paths[c_idx]
                
                record = {
                    'sheet': sheet_name,
                    'row': r_idx,
                    'col': c_idx,
                    'value': val,
                    'header_path': full_path,
                    'searchable_text': " ".join(full_path),
                    'type': self._get_value_type(val),
                    'table_name': table.get('name', f'Table{table_idx+1}')
                }
                self.records.append(record)

    
    def _build_column_paths(self, header_block: pd.DataFrame, 
                       style_grid: List[List[Dict]],
                       boundaries: Dict[str, int],
                       df: pd.DataFrame,
                       tables: List[Dict]) -> List[List[str]]:
    """Build hierarchical paths for each column with note filtering."""
    col_paths = []
    header_block_ffill = header_block.ffill(axis=1)
    
    for c in range(len(header_block.columns)):
        col_headers = header_block_ffill.iloc[:, c].tolist()
        
        clean_path = []
        for h_idx, h in enumerate(col_headers):
            if pd.isna(h):
                continue
                
            h_str = str(h).strip()
            if not h_str:
                continue
            
            # 🔥 NEW: Filter notes from headers
            actual_row_idx = header_block.index[h_idx]
            should_filter, reason = self.note_filter.should_filter_cell(
                h, actual_row_idx, c, boundaries, df, tables, 
                style_grid if style_grid else None
            )
            
            if should_filter:
                self.filtered_notes.append({
                    'sheet': '(current_sheet)',
                    'row': actual_row_idx,
                    'col': c,
                    'value': h,
                    'reason': f'header_{reason}'
                })
                continue  # Skip this header cell
            
            # Deduplicate consecutive identical headers
            if not clean_path or h_str != clean_path[-1]:
                clean_path.append(h_str)
        
        col_paths.append(clean_path)
    
    return col_paths

    def _build_row_path(self, row_ pd.Series, v_header_cols: List[int],
                       row_idx: int, boundaries: Dict[str, int],
                       df: pd.DataFrame, tables: List[Dict],
                       style_grid: List[List[Dict]],
                       sheet_name: str) -> List[str]:
        """Build hierarchical path from vertical headers with note filtering."""
        path = []
        
        for col_idx in v_header_cols:
            if col_idx >= len(row_data):
                continue
                
            val = row_data.iloc[col_idx]
            if pd.isna(val):
                continue
                
            val_str = str(val).strip()
            if not val_str:
                continue
            
            # 🔥 NEW: Filter notes from row headers
            should_filter, reason = self.note_filter.should_filter_cell(
                val, row_idx, col_idx, boundaries, df, tables,
                style_grid if style_grid else None
            )
            
            if should_filter:
                self.filtered_notes.append({
                    'sheet': sheet_name,
                    'row': row_idx,
                    'col': col_idx,
                    'value': val,
                    'reason': f'row_header_{reason}'
                })
                continue  # Skip this row header
            
            # Handle indentation for hierarchies
            val_clean = val_str.lstrip()
            indent_level = len(val_str) - len(val_clean)
            
            if indent_level > 0 and len(path) > 0:
                path = path[:-1]
            
            path.append(val_clean)
        
        return path

    
    def _build_embeddings(self, batch_size=100):
        """Build vector index."""
        unique_texts = list(set(r['searchable_text'] for r in self.records))
        
        for i in range(0, len(unique_texts), batch_size):
            batch = unique_texts[i:i+batch_size]
            try:
                vectors = self.matcher.embed_batch(batch)
                for text, vec in zip(batch, vectors):
                    self.vector_index[text] = vec
            except Exception as e:
                print(f"⚠️  Warning: Batch {i//batch_size + 1} embedding failed: {e}")

    def _get_value_type(self, val: Any) -> str:
        """Classify value type."""
        if isinstance(val, bool):
            return "boolean"
        if isinstance(val, (int, float)):
            if isinstance(val, float) and 0 <= abs(val) <= 1:
                return "percentage"
            return "number"
        if isinstance(val, str):
            if any(sym in val for sym in ['$', '€', '£', '¥']):
                return "currency"
            return "text"
        return "unknown"
    
    def _extract_critical_tokens(self, query: str) -> Set[str]:
        """Extract critical tokens from query."""
        tokens = set()
        query_lower = query.lower()
        parts = query_lower.split()
        
        for p in parts:
            clean_p = p.strip()
            if re.match(r'^(19|20)\d{2}$', clean_p):
                tokens.add(clean_p)
            elif re.match(r'^q[1-4]$', clean_p):
                tokens.add(clean_p)
            elif clean_p in ['january', 'february', 'march', 'april', 'may', 'june',
                           'july', 'august', 'september', 'october', 'november', 'december',
                           'jan', 'feb', 'mar', 'apr', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']:
                tokens.add(clean_p)
            elif clean_p in ['total', 'net', 'gross', 'operating', 'ebitda', 'ebit', 
                           'consolidated', 'adjusted', 'actual', 'budget', 'forecast']:
                tokens.add(clean_p)
        
        return tokens

    def query(self, question: str, top_k: int = 5, require_exact_tokens: bool = True) -> List[SearchResult]:
        """Precision query with hybrid scoring."""
        query_lower = question.lower()
        query_tokens = set(query_lower.split())
        critical_tokens = self._extract_critical_tokens(question)
        
        prefer_number = any(kw in query_lower for kw in 
                          ['how much', 'total', 'cost', 'revenue', 'profit', 'sum', 'amount', 'value'])
        
        scored_results = []
        q_vec = None
        
        if self.config.semantic_backend == 'openai':
            try:
                q_vec = self.matcher.embed_batch([question])[0]
            except Exception as e:
                print(f"⚠️  Warning: Query embedding failed: {e}")

        for r in self.records:
            target_text = r['searchable_text'].lower()
            
            is_exact = query_lower in target_text
            
            constraint_penalty = 1.0
            if critical_tokens:
                missing_critical = [t for t in critical_tokens if t not in target_text]
                if missing_critical:
                    constraint_penalty = 0.05 if require_exact_tokens else 0.4
            
            semantic_score = 0.0
            if self.config.semantic_backend == 'openai' and q_vec:
                target_vec = self.vector_index.get(r['searchable_text'])
                if target_vec:
                    semantic_score = self.matcher.cosine_similarity(q_vec, target_vec)
            else:
                semantic_score = self.matcher.calculate_similarity(question, r['searchable_text'])
            
            target_tokens = set(target_text.split())
            keyword_coverage = 0.0
            if query_tokens:
                intersection = query_tokens.intersection(target_tokens)
                keyword_coverage = len(intersection) / len(query_tokens)
            
            base_score = (semantic_score * 0.6) + (keyword_coverage * 0.4)
            
            if is_exact:
                base_score = min(0.99, base_score + self.config.exact_match_boost)
            
            final_score = base_score * constraint_penalty
            
            if prefer_number and r['type'] not in ['number', 'percentage', 'currency']:
                final_score *= 0.75
            
            if final_score >= self.config.min_confidence:
                scored_results.append((final_score, r, is_exact))
        
        scored_results.sort(key=lambda x: (not x[2], -x[0]))
        
        final_output = []
        for score, r, is_exact in scored_results[:top_k]:
            res = SearchResult(
                value=r['value'],
                confidence=score,
                sheet_name=r['sheet'],
                row=r['row'],
                col=r['col'],
                header_path=r['header_path'],
                context_text=r['searchable_text'],
                value_type=r['type'],
                is_exact_match=is_exact
            )
            final_output.append(res)
            
        return final_output
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get engine statistics."""
        sheets = set(r['sheet'] for r in self.records)
        tables = set((r['sheet'], r.get('table_name', '')) for r in self.records)
        
        # Group filtered notes by reason
        filter_reasons = {}
        for note in self.filtered_notes:
            reason = note['reason']
            filter_reasons[reason] = filter_reasons.get(reason, 0) + 1
        
        return {
            'total_records': len(self.records),
            'sheets': len(sheets),
            'tables': len(tables),
            'filtered_notes': len(self.filtered_notes),
            'filter_reasons': filter_reasons,
            'value_types': {vtype: sum(1 for r in self.records if r['type'] == vtype) 
                          for vtype in set(r['type'] for r in self.records)}
        }
    
    def get_filtered_notes(self, sheet_name: Optional[str] = None, 
                          reason: Optional[str] = None) -> List[Dict]:
        """Get filtered notes with optional filters."""
        notes = self.filtered_notes
        
        if sheet_name:
            notes = [n for n in notes if n['sheet'] == sheet_name]
        
        if reason:
            notes = [n for n in notes if n['reason'] == reason]
        
        return notes

# ==========================================
# 6. Example Usage
# ==========================================
if __name__ == "__main__":
    dummy_file = "financial_demo_with_notes.xlsx"
    
    if not os.path.exists(dummy_file):
        print("📝 Creating demo file with isolated notes...")
        
        with pd.ExcelWriter(dummy_file, engine='openpyxl') as writer:
            # Create P&L with isolated notes
            data = {
                'Line Item': ['Revenue', '  Product Sales', '  Service Revenue', 'Total Revenue',
                             '', 'Expenses', '  COGS', '  Operating Expenses', 'Total Expenses',
                             '', 'Net Income'] + [None]*5 + ['Note: All figures in millions'] + [None]*2,
                'Q1 2023': [100, 60, 40, 100, None, 70, 40, 30, 70, None, 30] + [None]*8,
                'Q2 2023': [120, 70, 50, 120, None, 80, 45, 35, 80, None, 40] + [None]*8,
                'Q3 2023': [110, 65, 45, 110, None, 75, 42, 33, 75, None, 35] + [None]*8,
                'Q4 2023': [130, 75, 55, 130, None, 85, 48, 37, 85, None, 45] + [None]*8
            }
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name='P&L Statement', index=False)

    api_key = os.getenv("OPENAI_API_KEY", "")
    
    config = QueryEngineConfig(
        semantic_backend='openai' if api_key else 'basic',
        openai_api_key=api_key if api_key else None,
        min_confidence=0.45,
        exact_match_boost=0.30,
        enable_footnote_filtering=True,
        aggressive_note_filtering=True,  # Enable aggressive filtering
        min_distance_from_data=2,
        isolation_window=3,
        sparse_region_threshold=0.2
    )

    try:
        print("\n" + "="*70)
        print("🚀 Financial Excel Engine - Universal Note Filtering")
        print("="*70 + "\n")
        
        engine = FinancialExcelEngine(dummy_file, config)
        
        stats = engine.get_statistics()
        print(f"\n📊 Engine Statistics:")
        print(f"   Total Records: {stats['total_records']}")
        print(f"   Filtered Notes: {stats['filtered_notes']}")
        print(f"   Filter Breakdown: {stats['filter_reasons']}")
        
        # Show filtered notes
        notes = engine.get_filtered_notes()
        if notes:
            print(f"\n🗑️  Filtered Notes/Annotations:")
            for note in notes[:10]:
                val_preview = str(note['value'])[:50]
                print(f"   - [{note['reason']}] R{note['row']+1}C{note['col']+1}: {val_preview}")
        
        print("\n" + "="*70)
        print("🔎 Running Test Queries")
        print("="*70)
        
        test_queries = [
            "Revenue Q1 2023",
            "Net Income Q4 2023",
            "Service Revenue Q2 2023",
        ]
        
        for query in test_queries:
            print(f"\n📌 Query: '{query}'")
            results = engine.query(query, top_k=3)
            
            if not results:
                print("   ❌ No results found")
            else:
                for idx, r in enumerate(results, 1):
                    exact = "🎯 EXACT" if r.is_exact_match else ""
                    path = " > ".join(r.header_path)
                    print(f"   [{idx}] {exact}")
                    print(f"       Value: {r.value} ({r.value_type})")
                    print(f"       Confidence: {r.confidence:.3f}")
                    print(f"       Path: {path}")
                
    except Exception as e:
        print(f"\n❌ Error: {e}")
